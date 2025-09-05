from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import json

# very much hack
# look for company legal company
# find the income statement
# print the following tables until statement of equity
# will break under following conditions:
# - income statmenet is not first
# - statement of equity is not first
# - income statement is on two or more pages
# - statement of equity is on two or more pages

def main():
    # apple 10q --> aapl-20250628
    fileToParsePath = "../appl.htm"
    # TODO: will need to search for
    fileToParsePathD = "/mnt/c/Users/Michi/Documents/codeSpace/appl.htm"

    # open the file
    fileToParse = open(fileToParsePathD)
    soupDiv = (BeautifulSoup(fileToParse, "html.parser")).find_all("div")

    # TODO: will need to search for later
    keyWord = "Apple Inc."
    # flag for when we want to search the following div
    hitFlag = 0 
    extractTable = False

    for divToSearch in soupDiv:
        
        if extractTable:
             finTable = divToSearch.find("table")
             if finTable != None:
                  printToNewExcelSheet(finTable)

        divStrings = divToSearch.strings
        for stringToSearch in divStrings:
            hitFlag, extractTable = matchFlagger(stringToSearch, keyWord, hitFlag, extractTable)
        
             


def matchFlagger(stringToMach, keyWord, hitFlag, boolFlag):
    # max amount of lines to search 
    maxAttempts = 2
    # TODO: fix so program can grab all 4 statements
    titlesOfInterest = ["STATEMENTS OF OPERATIONS", "STATEMENTS OF CASH FLOWS"]
    titleToMatch = titlesOfInterest[0] if not boolFlag else titlesOfInterest[1]

    # look for keyword to flag to search the next line
    if hitFlag == 0:
            if keyWord in stringToMach:
                hitFlag += 1

    # search next line for title of interest
    # if value exists, switch state of boolFlag
    # if searchflag exceeds maxAttempts stop searching the lines, else try next line
    # TODO: implement array for title of interest
    else:
            if titleToMatch in stringToMach:
                boolFlag = True if not boolFlag else False
                hitFlag = 0
            hitFlag = 0 if hitFlag > maxAttempts else + 1

    # returns hitFlag value and boolFlag
    return hitFlag, boolFlag

# WARNING: this shit reeks
def printToNewExcelSheet(finTable):
    ftotal = "/mnt/c/Users/Michi/Desktop/tets.xlsx"
    xlTemplate = load_workbook(filename = ftotal)
    ws1 = xlTemplate.create_sheet("FINSTATEMENT")

    # TODO: have this shit print into an array for easier manipulation
    # TODO: turn finstatements in objects?
    finTableRows = finTable.find_all("tr")

    x_axis = 0
    for tableRow in finTableRows:
         tableData = tableRow.find_all("td")
         y_axis = 1
         x_axis += 1
         for data in tableData:
             dataString = data.strings
             for stringData in dataString:
                  ws1.cell(row = x_axis, column = y_axis, value = stringData)
                  y_axis += 1
        
    
    
    xlTemplate.save(ftotal)



main()
