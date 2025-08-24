# TODO:
"""
- open template JSON file
- open excel file
- Save Value keys to an array
- Soupify HTML and chop it into its table parts by the row
OR
- Run imaging api to save values by the row
- save rows to a dict, first coloumn being the key and the remaining being the values
- make copy of excel template
- load dict to excel copy via pandas 
- save and close excel copy and excel template
"""

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import json

def main():
    # apple 10q --> aapl-20250628
    fileToParsePath = "../appl.htm"
    fileToParsePathD = "/mnt/c/Users/Michi/Documents/codeSpace/appl.htm"

    # open the file
    fileToParse = open(fileToParsePathD)

    # print out contents thereof
    soup = (BeautifulSoup(fileToParse, "html.parser"))
     #parse finSynDict and load sudo array

    # put html Table rows in dicts

    tableTree = soup.find_all("table")
    for tTable in tableTree:
        treeRows = tTable.find_all("tr")
        tableContents = {}
        tableHasher = []
        for treeRow in treeRows:
            treeRowStringsArr = list(treeRow.strings)
            if(len(treeRowStringsArr) > 0):
                tableHasher.append(treeRowStringsArr[0])
                tableContents[treeRowStringsArr[0]] = treeRowStringsArr[1:]
        test(tableHasher, tableContents)
        # verify function with JSON stuff 

    """
    # print to excel --> prob wont work
    ftotal = "/mnt/c/Users/Michi/Desktop/tets.xlsx"
    xlTemplate = load_workbook(filename = ftotal)
    ws1 = xlTemplate.create_sheet("Balance Sheet")
    x = 1
    for i in range(len(tableHasher)):
        rowArray = tableContents[tableHasher[i]]
        y = 1
        ws1.cell(row = x, column = y, value = tableHasher[i])
        for j in range(len(rowArray)):
            y += 1
            ws1.cell(row = x, column = y, value = rowArray[j])
        x += 1
    xlTemplate.save(ftotal)
    """

def test(hash, contents):
    with open('/mnt/c/Users/Michi/Documents/codeSpace/excel_template_autofill_python/finSynDict.json') as finDictRaw:
        finDict = json.load(finDictRaw)

    
    for tableName in finDict:
        tableValueArr = [] # will contain the arrays of each fin statement
        for tableRowNames in finDict[tableName]:
            tableValueArr.append(finDict[tableName][tableRowNames])
        # here we can make operations on the table's arrays
        tablePossibleValueArr = [] # contains all possible values of the tables coloumns
        for i in range (len(tableValueArr)):
            for j in range (len(tableValueArr[i])):
                tablePossibleValueArr.append(tableValueArr[i][j])
        finTableConf = arrComp(hash, tablePossibleValueArr)
        if finTableConf > 50:
            print(tableName)
            print(contents)
        
def arrComp(hash, compArr):
    n = 0
    success = 0
    lcompArr = [s.lower() for s in compArr]
    for hashValue in hash:
        n += 1
        if hashValue.lower() in lcompArr:
            success += 1
    if n > 0:
        return (success/n)*100
    else: 
        return n



main()

